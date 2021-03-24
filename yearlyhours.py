import os

import openpyxl
from datetime import datetime

file_name = None
mytasks = []
myyears = []
daysinyear = []
weeksinyear = []

daterow = str(2)
taskrow = 5  # row that task codes start at
taskcolumn = 'D'

with open('IgnoredTaskCodes.txt', 'r') as f:  # opening this way closes the file when finished, so no f.close required
    ignoredtasks = f.read().splitlines()

ignoredtasks.sort()
print('Task Codes to Ignore:')
print(ignoredtasks)

cwd = os.getcwd()
print(cwd)

print("Loading workbook")
print('TimeSheetReportYearly.xlsx')
# quick-load a temporary workbook
wbtemp = openpyxl.load_workbook('TimeSheetReportYearly.xlsx', read_only=True)

mysheets = wbtemp.sheetnames
sheet_total = len(mysheets)
sheet_count = 0


# print out a loading percentage as sheets are loaded.
print('Loaded: {:.0%}'.format(sheet_count / sheet_total))

# create workbook that will be used in rest of program
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove initial sheet

for sheet in mysheets:
    sheet_count += 1
    wb.create_sheet(sheet)
    ws = wb[sheet]
    wstemp = wbtemp[sheet]

    # iterate through rows in temp sheet and copy the cells into the working sheet
    row_count = 1
    for row in wstemp.values:
        column_count = 1
        for value in row:
            ws.cell(row=row_count, column=column_count, value=value)
            column_count += 1
        row_count += 1

    print('Loaded: {:.0%}'.format(sheet_count / sheet_total))


hourcolls = [[] for i in mysheets]

# get tasks from timesheet report
for i in range(len(mysheets)):
    sheet = wb.worksheets[i]
    sheetname = wb.sheetnames[i]

    # create hour column for each sheet
    print("Finding hour columns for sheet: " + sheetname)
    for m in range(1, sheet.max_column-2):
        if sheet.cell(row=4, column=m).value == "Use":
            colname = openpyxl.utils.cell.get_column_letter(m)
            hourcolls[i].append(colname)

    print("Finding tasks and years for sheet: " + sheetname)
    j = taskrow
    while True:
        # Task Code column is hard coded
        cell = taskcolumn + str(j)
        task = sheet[cell].value
        if task is None:
            for k in range(len(hourcolls[i])):
                datecell = hourcolls[i][k] + daterow
                mydate = sheet[datecell].value
                if type(mydate) is datetime:
                    mydate = mydate.strftime("%d/%m/%Y")
                year = mydate[-2:]
                if year not in myyears:
                    myyears.append(year)
            break
        if task not in mytasks:
            mytasks.append(task)
        j += 1

mytasks.sort()
print('\r\nFound the Following Task Codes:')
print(mytasks)

myyears.sort()
print('\r\nFound the Following Years:')
print(myyears)

# get hours
taskhours = [0] * len(mytasks)                                # this may be able to be deleted - investigate
myhours = [[0] * len(mytasks) for i in myyears]
daysinyear = [0] * len(myyears)
weeksinyear = [0] * len(myyears)

for i in range(len(mysheets)):
    sheet = wb.worksheets[i]
    sheetname = wb.sheetnames[i]
    print("Finding task hours for sheet: " + sheetname)
    j = taskrow
    while True:
        taskcell = taskcolumn + str(j)
        task = sheet[taskcell].value
        if task is None:
            break
        taskindex = mytasks.index(task)
        for k in range(len(hourcolls[i])):
            datecell = hourcolls[i][k] + daterow
            mydate = sheet[datecell].value
            if type(mydate) is datetime:
                mydate = mydate.strftime("%d/%m/%Y")
            year = mydate[-2:]
            yearindex = myyears.index(year)
            if j == taskrow:
                daysinyear[yearindex] += 1
            hourcell = hourcolls[i][k] + str(j)
            if sheet[hourcell].value is not None:
                myhours[yearindex][taskindex] += sheet[hourcell].value

        j += 1

totalhours = 0
ignoredhours = 0
totalbyyear = [0] * len(myyears)
ignoredbyyear = [0] * len(myyears)
workedbyyear = [0] * len(myyears)
print('\r\nTask Code Hours')
for i in range(len(mytasks)):
    for j in range(len(myyears)):
        taskhours[i] += myhours[j][i]
        totalbyyear[j] += myhours[j][i]
    totalhours += taskhours[i]
    if mytasks[i] in ignoredtasks:
        ignoredhours += taskhours[i]
        for j in range(len(myyears)):
            ignoredbyyear[j] += myhours[j][i]
    print(mytasks[i] + ' ' + str(taskhours[i]))

weeksworked = 0
for i in range(len(myyears)):
    weeksinyear[i] = round(daysinyear[i]/7)
    weeksworked += weeksinyear[i]


def hours_format(hours):
    if hours >= 10000:
        formatted = ' ' + '{:.2f}'.format(hours)
    elif hours >= 1000:
        formatted = '  ' + '{:.2f}'.format(hours)
    elif hours >= 100:
        formatted = '   ' + '{:.2f}'.format(hours)
    elif hours >= 10:
        formatted = '    ' + '{:.2f}'.format(hours)
    else:
        formatted = '     ' + '{:.2f}'.format(hours)
    return formatted


def weeks_format(weeks):
    if weeks >= 10000:
        formatted = '  ' + str(weeks) + '  '
    elif weeks >= 1000:
        formatted = '   ' + str(weeks) + '  '
    elif weeks >= 100:
        formatted = '    ' + str(weeks) + '  '
    elif weeks >= 10:
        formatted = '     ' + str(weeks) + '  '
    else:
        formatted = '      ' + str(weeks) + '  '
    return formatted


workedhours = totalhours - ignoredhours

totalheader = '                Year:'
totalline = 'Total Hours:  ' + hours_format(totalhours)
ignoredline = 'Ignored Hours:' + hours_format(ignoredhours)
workedline = 'Working Hours:' + hours_format(workedhours)
weeksline = 'Weeks Worked: ' + weeks_format(weeksworked)
averageline = 'Hours / Week: ' + hours_format(workedhours / weeksworked)

for i in range(len(myyears)):
    totalheader += '       ' + myyears[i]
    totalline += hours_format(totalbyyear[i])
    ignoredline += hours_format(ignoredbyyear[i])
    workedline += hours_format(totalbyyear[i] - ignoredbyyear[i])
    weeksline += weeks_format(weeksinyear[i])
    if weeksinyear[i] > 0:
        averageline += hours_format((totalbyyear[i] - ignoredbyyear[i]) / weeksinyear[i])

print('weeks in year')
print(weeksinyear)

print('\r\n' + totalheader)
print(totalline)
print(ignoredline)
print(workedline)
print(weeksline)
print(averageline)

print('\r\n')
action = "Z"
while action != "N":
    action = input("Do you want to calculate $/hr? [Y/N] ").upper()
    if action not in "YN" or len(action) != 1:
        print("Please choose [Y]es or [N]o")
        continue
    if action == "Y":
        break
    else:
        exit()

myyearlypay = [0] * len(myyears)

rerun = "Z"
while rerun != "N":
    input_action = "Z"
    while input_action != "N":
        input_action = input("Would you like to [I]mport your yearly pay or [M]anually enter it? ").upper()
        if input_action not in "IM" or len(action) != 1:
            print("Please choose [I]port or [M]anual entry")
            continue
        if input_action == "I":
            file_action = "Z"
            while file_action != "N":
                file_action = input("Use default file (YearlyPayImport.txt)? [Y/N] ").upper()
                if file_action not in "YN" or len(action) != 1:
                    print("Please choose [Y]es or [N]o")
                    continue
                if file_action == "Y":
                    if os.path.isfile('YearlyPayImport.txt'):  # consider adding " and os.access(PATH, os.R_OK)"
                        print('Importing file ' + cwd + '\\YearlyPayImport.txt')
                        file_name = "YearlyPayImport.txt"
                        break
                    else:
                        print('Import file doesn\'t exist, creating file: ' + cwd + '\\YearlyPayImport.txt')
                        with open('YearlyPayImport.txt', 'w') as f:
                            f.write('Enter each year on a new line as Year,Pay (EX 2000,50000)')
                            f.write('')
                        input('Edit the Import file, then press Enter.')
                if file_action == "N":
                    file_name = input("Enter name of file to use: ")
                    if file_name[-4:].upper() != ".TXT":
                        file_name = file_name + ".txt"
                    if os.path.isfile(file_name):
                        break
                    else:
                        print("File not found")
                        # reset action to allow while loop to ask about file again
                        file_action = "Z"

            with open(file_name, 'r') as f:
                next(f)
                for line in f:
                    yearstr, paystr = line.split(',')
                    year = yearstr[-2:]
                    if year not in myyears:
                        print('No matching year found: ' + yearstr)
                    else:
                        yearindex = myyears.index(year)
                        myyearlypay[yearindex] = int(paystr)
            break
        else:
            print('Enter how much you were paid each year (enter 0 to skip that year)')
            for i in range(len(myyears)):
                myyearlypay[i] = int(input('How much were you paid in \'' + myyears[i] + '? $').replace(',', ''))
            break

    adjustedtotalhours = 0
    adjustedignoredhours = 0
    for i in range(len(myyears)):
        if myyearlypay[i] != 0:
            adjustedtotalhours += totalbyyear[i]
            adjustedignoredhours += ignoredbyyear[i]

    perhourheader = '                Year:'
    perhourtotalline = 'Total $/hr:   ' + hours_format(sum(myyearlypay) / adjustedtotalhours)
    perhourworkedline = 'Worked $/hr:  ' + hours_format(sum(myyearlypay) / (adjustedtotalhours - adjustedignoredhours))

    for i in range(len(myyears)):
        perhourheader += '       ' + myyears[i]
        if myyearlypay[i] == 0:
            perhourtotalline += '         '
            perhourworkedline += '         '
        else:
            perhourtotalline += hours_format(myyearlypay[i] / totalbyyear[i])
            perhourworkedline += hours_format(myyearlypay[i] / (totalbyyear[i] - ignoredbyyear[i]))

    print('\r\n' + perhourheader)
    print(perhourtotalline)
    print(perhourworkedline)

    print('\r\nTo re-import your TimeSheet, restart the application')

    rerun = "Z"
    while rerun not in "YN":
        rerun = input("Would you like to import a different pay file? [Y/N] ").upper()
