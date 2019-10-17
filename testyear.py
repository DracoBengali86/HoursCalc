# import openpyxl
import os
# from openpyxl import load_workbook
# from openpyxl import Workbook
import openpyxl
from datetime import datetime

mytasks = []
ignoredtasks = []
myyears = []
daysinyear = []
weeksinyear = []

daterow = str(2)
taskrow = 5  # row that task codes start at
taskcolumn = 'D'

with open('IgnoredTaskCodes.txt', 'r') as f:  #opening this way closes the file when finished, so no f.close required
    ignoredtasks = f.read().splitlines()

ignoredtasks.sort()
print('Task Codes to Ignore:')
print(ignoredtasks)

cwd = os.getcwd()
print(cwd)
# os.chdir("/path/to/folder")
# os.listdir('.')

print("Loading workbook (this takes time, please be patient)")
wb = openpyxl.load_workbook('TimeSheetReportYearly.xlsx')
# print(wb2.sheetnames)
# mylength = len(wb2.sheetnames)
# print(mylength)

mysheets = wb.sheetnames

hourcolls = [[] for i in mysheets]

# get tasks from timesheet report
for i in range(len(mysheets)):
    sheet = wb.worksheets[i]
    sheetname = wb.sheetnames[i]

    # create hour column for each sheet
    print("Finding hour columns for sheet: " + sheetname)
    for m in range(1,sheet.max_column-2):
        if sheet.cell(row=4,column=m).value == "Use":
            colname = openpyxl.utils.cell.get_column_letter(m)
            hourcolls[i].append(colname)

    print("Finding tasks and years for sheet: " + sheetname)
    j = taskrow
    while True:
        # Task Code column is hard coded
        cell = taskcolumn + str(j)
        task = sheet[cell].value
        if task is None:
            for k in range(len(hourcolls[i])):
                datecell = hourcolls[i][k] + daterow
                mydate = sheet[datecell].value
                if type(mydate) is datetime:
                    mydate = mydate.strftime("%d/%m/%Y")
                year = mydate[-2:]
                if year not in myyears:
                    myyears.append(year)
            break
        if task not in mytasks:
            mytasks.append(task)
        j += 1

mytasks.sort()
print('\r\nFound the Following Task Codes:')
print(mytasks)

myyears.sort()
print('\r\nFound the Following Years:')
print(myyears)

# get hours
taskhours = [0] * len(mytasks)                                  #this may be able to be deleted - investigate
myhours = [[0] * len(mytasks) for i in myyears]
daysinyear = [0] * len(myyears)
weeksinyear = [0] * len(myyears)

#def printhours():
#    for row in myhours:
#        for elem in row:
#            print(elem, end=' ')
#        print()

#printhours()

for i in range(len(mysheets)):
    sheet = wb.worksheets[i]
    sheetname = wb.sheetnames[i]
    print("Finding task hours for sheet: " + sheetname)
    j = taskrow
    while True:
        taskcell = taskcolumn + str(j)
        task = sheet[taskcell].value
        if task is None:
            break
        taskindex = mytasks.index(task)
        for k in range(len(hourcolls[i])):
            datecell = hourcolls[i][k] + daterow
            mydate = sheet[datecell].value
            if type(mydate) is datetime:
                mydate = mydate.strftime("%d/%m/%Y")
            year = mydate[-2:]
            yearindex = myyears.index(year)
            if j == taskrow:
                daysinyear[yearindex] += 1
#            print(mydate)
#            if year not in myyears:
#                myyears.append(year)
            hourcell = hourcolls[i][k] + str(j)
#            print(sheet[hourcell].value)
            if sheet[hourcell].value != None:
#                taskhours[taskindex] += sheet[hourcell].value
                myhours[yearindex][taskindex] += sheet[hourcell].value
#                print('partial array')
#                printhours()

        j += 1

#print('filled array')
#printhours()

#print('days in year')
#print(daysinyear)

totalhours = 0
ignoredhours = 0
totalbyyear = [0] * len(myyears)
ignoredbyyear = [0] * len(myyears)
workedbyyear = [0] * len(myyears)
print('\r\nTask Code Hours')
for i in range(len(mytasks)):
    for j in range(len(myyears)):
        taskhours[i] += myhours[j][i]
        totalbyyear[j] += myhours[j][i]
    totalhours += taskhours[i]
    if mytasks[i] in ignoredtasks:
        ignoredhours += taskhours[i]
        for j in range(len(myyears)):
            ignoredbyyear[j] += myhours[j][i]
    print(mytasks[i] + ' ' + str(taskhours[i]))

weeksworked = 0
for i in range(len(myyears)):
    weeksinyear[i] = round(daysinyear[i]/7)
    weeksworked += weeksinyear[i]

def hoursformat(hours):
    if hours >= 10000:
        formatted = ' ' + '{:.2f}'.format(hours)
    elif hours >= 1000:
        formatted = '  ' + '{:.2f}'.format(hours)
    elif hours >= 100:
        formatted = '   ' + '{:.2f}'.format(hours)
    elif hours >= 10:
        formatted = '    ' + '{:.2f}'.format(hours)
    else:
        formatted = '     ' + '{:.2f}'.format(hours)
    return formatted


def weeksformat(weeks):
    if weeks >= 10000:
        formatted = '  ' + str(weeks) + '  '
    elif weeks >= 1000:
        formatted = '   ' + str(weeks) + '  '
    elif weeks >= 100:
        formatted = '    ' + str(weeks) + '  '
    elif weeks >= 10:
        formatted = '     ' + str(weeks) + '  '
    else:
        formatted = '      ' + str(weeks) + '  '
    return formatted


workedhours = totalhours - ignoredhours

totalheader = '                Year:'
totalline = 'Total Hours:  ' + hoursformat(totalhours)
ignoredline = 'Ignored Hours:' + hoursformat(ignoredhours)
workedline = 'Working Hours:' + hoursformat(workedhours)
weeksline = 'Weeks Worked: ' + weeksformat(weeksworked)
# averageline = 'Hours / Week: ' + hoursformat(workedhours/(52*len(myyears)))
averageline = 'Hours / Week: ' + hoursformat(workedhours/weeksworked)

for i in range(len(myyears)):
    totalheader += '       ' + myyears[i]
    totalline += hoursformat(totalbyyear[i])
    ignoredline += hoursformat(ignoredbyyear[i])
    workedline += hoursformat(totalbyyear[i] - ignoredbyyear[i])
    weeksline += weeksformat(weeksinyear[i])
    if weeksinyear[i] > 0:
        averageline += hoursformat((totalbyyear[i] - ignoredbyyear[i])/weeksinyear[i])

print('weeks in year')
print(weeksinyear)

print('\r\n' + totalheader)
print(totalline)
print(ignoredline)
print(workedline)
print(weeksline)
print(averageline)

print('\r\n')
action = "Z"
while action != "N":
    action = input("Do you want to calculate $/hr? [Y/N]").upper()
    if action not in "YN" or len(action) != 1:
        print("Please choose [Y]es or [N]o")
        continue
    if action == "Y":
        #mypay = int(input("How much were you paid? $"))
        #print('\r\nTotal $/hr:  ' + '{:.2f}'.format(mypay/totalhours))
        #print('Worked $/hr: ' + '{:.2f}'.format(mypay/workedhours))
        break
    else:
        exit()

myyearlypay = [0] * len(myyears)

action = "Z"
while action != "N":
    action = input("Would you like to import your yearly pay? [Y/N]").upper()
    if action not in "YN" or len(action) != 1:
        print("Please choose [Y]es or [N]o")
        continue
    if action == "Y":
        if os.path.isfile('YearlyPayImport.txt'):  #consider adding " and os.access(PATH, os.R_OK)"
            print('Importing file ' + cwd + '\\YearlyPayImport.txt')
        else:
            print('Import file doesn\'t exist, creating file: ' + cwd + '\\YearlyPayImport.txt')
            with open('YearlyPayImport.txt', 'w') as f:
                f.write('Enter each year on a new line as Year,Pay (EX 2000,50000)')
                f.write('')
            input('Edit the Import file, then press Enter.')
        with open('YearlyPayImport.txt', 'r') as f:
            next(f)
            for line in f:
                yearstr, paystr = line.split(',')
                year = yearstr[-2:]
                if year not in myyears:
                    print('No matching year found: ' + yearstr)
                else:
                    yearindex = myyears.index(year)
                    myyearlypay[yearindex] = int(paystr)
        break
    else:
        print('Enter how much you were paid each year (enter 0 to skip)')
        for i in range(len(myyears)):
            myyearlypay[i] = int(input('How much were you paid in \'' + myyears[i] + '? $'))
        break

adjustedtotalhours = 0
adjustedignoredhours = 0
for i in range(len(myyears)):
    if myyearlypay[i] != 0:
        adjustedtotalhours += totalbyyear[i]
        adjustedignoredhours += ignoredbyyear[i]

perhourheader = '                Year:'
perhourtotalline = 'Total $/hr:   ' + hoursformat(sum(myyearlypay) / adjustedtotalhours)
perhourworkedline = 'Worked $/hr:  ' + hoursformat(sum(myyearlypay) / (adjustedtotalhours - adjustedignoredhours))

for i in range(len(myyears)):
    perhourheader += '       ' + myyears[i]
    if myyearlypay[i] == 0:
        perhourtotalline += '         '
        perhourworkedline += '         '
    else:
        perhourtotalline += hoursformat(myyearlypay[i] / totalbyyear[i])
        perhourworkedline += hoursformat(myyearlypay[i] / (totalbyyear[i] - ignoredbyyear[i]))

print('\r\n' + perhourheader)
print(perhourtotalline)
print(perhourworkedline)

#work on getting pay enterable from text file
#maybe csv with year,pay (ignore $ sign? or require it not to be there?)
#could just loop with every other line being year/pay
#ask if user wants to import, if file doesn't exist create blank with "example" entry
#verify/wait for user to say they've filled it in