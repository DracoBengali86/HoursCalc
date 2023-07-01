import os
import pandas as pd
import time

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from datetime import date, datetime, timedelta

# TODO: Create function to create string from datetime in d/m/yy with no zero padding
# TODO: Create function to find previous Sunday and return formatted date string (possibly make selectable day)
today = datetime.now()
year = str(today.year)
offset = today.isoweekday()
sunday = today - timedelta(days=offset)
new_year_test = date(int(year), 1, 15)
is_new_year = False

if today.date() < new_year_test:
    print("Start of New Year")
    is_new_year = True
    year = str(int(year)-1)

# Start date must be formatted d/m/yy, not dd/mm/yyyy it is used to compare to the date on the spreadsheet
# and that date has no zero padding and only a two digit year. End date only has to exclude time info (no HH:MM:SS.SS)
start_date = '1/1/' + year[-2:]

if is_new_year:
    end_date = '12/31/' + year[-2:]
else:
    end_date = sunday.strftime("%m/%d/%y")

key_control = u'\ue009'
key_backspace = u'\ue003'
key_delete = u'\ue017'
key_tab = u'\ue004'


def main():
    driver = webdriver.Chrome()
    driver.get("https://apps.tricore.com/MobileTime/rep-timesheet")

    url = driver.current_url

    print(driver.page_source)
    if "adfs.tricore.com/adfs" in url:
        action = input("Is login complete? [Y/N]").upper()
        if action == "N":
            exit()
        url = driver.current_url

    if "rep-timesheet" not in url:
        driver.get("https://apps.tricore.com/MobileTime/rep-timesheet")

    try:
        element_present = ec.presence_of_element_located((By.ID, 'startDateInput'))
        WebDriverWait(driver, 60).until(element_present)
    except TimeoutException:
        print("Timed out waiting for page load")
        driver.quit()
        exit(20)

    time.sleep(2)
    start_element = driver.find_element("id", "startDateInput")
    end_element = driver.find_element("id", "endDateInput")
    # clear End Date field
    end_element.send_keys(u'\ue009' + "a" + u'\ue003')
    # enter new end date
    end_element.send_keys(end_date)
    # clear Start Date field (Ctrl + a + backspace)
    start_element.send_keys(u'\ue009' + "a" + u'\ue003')
    # enter new start date
    start_element.send_keys(start_date)
    # tab away to get page to update
    end_element.send_keys("\t")

    # try to wait for page load
    try:
        element = ec.text_to_be_present_in_element((By.XPATH, "//th[2]"), start_date)
        WebDriverWait(driver, 10).until(element)
    except TimeoutException:
        print("Timed out waiting for page load")
        driver.quit()
        exit(21)

    df = pd.read_html(driver.page_source)[0]
    print("Head\n")
    print(df.head())
    print("\n\nNon Head")
    print(df)
    df.to_excel(r'temp.xlsx', header=True)

    driver.quit()

    # modify_spreadsheet(df)
    wb = openpyxl.load_workbook('temp.xlsx')
    print("Building Spreadsheet")
    build_spreadsheet(wb)
    print("Removing temp files")
    os.remove('temp.xlsx')

    exit()


def build_spreadsheet(workbook):
    data_ws = workbook.active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test"

    project_count = 0
    hours_count = 0
    single = 12.0
    double = 24.0

    df = Font(name='Segoe UI', size=8, bold=False)
    dfb = Font(name='Segoe UI', size=8, bold=True)

    da = Alignment(wrap_text=True, vertical='center')
    dac = Alignment(wrap_text=True, vertical='center', horizontal='center')
    dar = Alignment(wrap_text=True, vertical='center', horizontal='right')
    atc = Alignment(wrap_text=True, vertical='top', horizontal='center')

    dbd = PatternFill(start_color='48606D', fill_type='solid')
    dbm = PatternFill(start_color='628395', fill_type='solid')
    dbl = PatternFill(start_color='D3D3D3', fill_type='solid')

    dbs = Side(style='medium', color='A9A9A9')
    dbn = Side(style='none')

    ws.sheet_format.defaultColWidth = 6.7109375

    ws.row_dimensions[1].height = single
    ws.row_dimensions[2].height = single
    ws.row_dimensions[3].height = single
    ws.row_dimensions[4].height = double
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 25.5703125
    ws.column_dimensions['C'].width = 24.42578125
    ws.column_dimensions['D'].width = 25.0

    ws.merge_cells('A2:E3')
    ws['B2'].border = Border(top=dbs)
    ws['C2'].border = Border(top=dbs)
    ws['D2'].border = Border(top=dbs)
    ws['E2'].border = Border(top=dbs, right=dbs)
    ws['E3'].border = Border(right=dbs)
    format_cell(ws, 'A2', src_cell='B2', src_ws=data_ws, alignment=dac, font=dfb, bg_color=dbd)
    format_cell(ws, 'A4', text='Project', alignment=dac, font=df)
    format_cell(ws, 'B4', text='Project Name', alignment=dac, font=df)
    format_cell(ws, 'C4', text='Account', alignment=dac, font=df)
    format_cell(ws, 'D4', text='Task Code', alignment=dac, font=df)
    format_cell(ws, 'E4', text='Sprint Cat.', alignment=dac, font=df)

    taskcolumn = "E"
    taskrow = 5
    while True:
        cell = taskcolumn + str(taskrow)
        task = data_ws[cell].value
        if task is None:
            break
        dest_col = ['A', 'B', 'C', 'D', 'E']
        src_col = ['B', 'C', 'D', 'E', 'F']
        for i in range(len(dest_col)):
            dest = dest_col[i]+str(taskrow)
            src = src_col[i]+str(taskrow)
            format_cell(ws, dest, src_cell=src, src_ws=data_ws, alignment=da, font=df, height=double)
            i += 1
        project_count += 1
        taskrow += 1

    row = 'E' + str(project_count+5)
    format_cell(ws, row, text='Totals:', alignment=dar, font=dfb, bg_color=dbm, height=single)

    hour_column = 7
    hour_row = 1
    while True:
        date_cell = get_column_letter(hour_column) + str(hour_row)
        day_cell = get_column_letter(hour_column) + str(hour_row + 1)
        hours = data_ws[date_cell].value
        if hours is None:
            break
        # Date cell
        dest = get_column_letter(hour_column - 1) + str(hour_row + 1)
        merge = dest + ':' + get_column_letter(hour_column) + str(hour_row + 1)
        ws.merge_cells(merge)
        format_cell(ws, dest, src_cell=date_cell, src_ws=data_ws, alignment=dac, font=dfb, bg_color=dbm)
        ws[dest].border = Border(top=dbs, bottom=dbn)
        ws[get_column_letter(hour_column) + str(hour_row + 1)].border = Border(top=dbs, right=dbs)
        ws[dest].number_format = 'DD/MM/YYYY'

        # Day of Week Cell
        dest = get_column_letter(hour_column - 1) + str(hour_row + 2)
        merge = dest + ':' + get_column_letter(hour_column) + str(hour_row + 2)
        ws.merge_cells(merge)
        format_cell(ws, dest, src_cell=day_cell, src_ws=data_ws, alignment=dac, font=dfb, bg_color=dbm)
        ws[dest].border = Border(top=dbn)
        ws[get_column_letter(hour_column) + str(hour_row + 2)].border = Border(right=dbs)

        # Use/Chr cells
        dest = get_column_letter(hour_column - 1) + str(hour_row + 3)
        format_cell(ws, dest, text="Use", alignment=dac, font=dfb)
        dest = get_column_letter(hour_column) + str(hour_row + 3)
        format_cell(ws, dest, text="Chg", alignment=dac, font=dfb, bg_color=dbl)

        hour_column += 2
        hours_count += 2

    date_dest = get_column_letter(hours_count + 4) + '2'
    day_dest = get_column_letter(hours_count + 4) + '3'
    ws.unmerge_cells(date_dest + ':' + get_column_letter(hours_count + 5) + '2')
    ws.unmerge_cells(day_dest + ':' + get_column_letter(hours_count + 5) + '3')
    merge = date_dest + ':' + get_column_letter(hours_count+5) + '3'
    ws.merge_cells(merge)
    ws.cell(row=2, column=hours_count + 5).border = Border(top=dbs, right=dbs)
    ws.cell(row=3, column=hours_count + 5).border = Border(right=dbs, bottom=dbs)
    ws.cell(row=3, column=hours_count + 4).border = Border(bottom=dbs)
    format_cell(ws, date_dest, text="Totals", alignment=atc, font=dfb, bg_color=dbm)
    dest = get_column_letter(hours_count + 4) + '4'
    format_cell(ws, dest, text="Use", alignment=dac, font=dfb, bg_color=dbm)
    dest = get_column_letter(hours_count + 5) + '4'
    format_cell(ws, dest, text="Chg", alignment=dac, font=dfb, bg_color=dbm)

    for i in range(project_count + 1):
        if i == project_count:
            bgc1 = dbm
            bgc2 = dbm
            ft = dfb
        else:
            bgc1 = None
            bgc2 = dbl
            ft = df
        for j in range(0, hours_count - 2, 2):
            # Use column
            dest = get_column_letter(j + 6) + str(i + 5)
            src = get_column_letter(j + 7) + str(i + 5)
            format_cell(ws, dest, src_cell=src, src_ws=data_ws, alignment=dac, bg_color=bgc1, font=ft)
            ws[dest].number_format = '0.00#'

            # Chr column
            dest = get_column_letter(j + 7) + str(i + 5)
            src = get_column_letter(j + 8) + str(i + 5)
            format_cell(ws, dest, src_cell=src, src_ws=data_ws, alignment=dac, bg_color=bgc2, font=ft)
            ws[dest].number_format = '0.00#'

        # Used column
        dest = get_column_letter(hours_count + 4) + str(i + 5)
        src = get_column_letter(hours_count + 5) + str(i + 5)
        format_cell(ws, dest, src_cell=src, src_ws=data_ws, alignment=dac, bg_color=dbm, font=ft)
        ws[dest].number_format = '0.00#'

        # Charged column
        dest = get_column_letter(hours_count + 5) + str(i + 5)
        src = get_column_letter(hours_count + 6) + str(i + 5)
        format_cell(ws, dest, src_cell=src, src_ws=data_ws, alignment=dac, bg_color=dbm, font=ft)
        ws[dest].number_format = '0.00#'

    wb.save('TimeSheetReportCurrentYear.xlsx')


def format_cell(ws, dest_cell, alignment=None, font=None, bg_color=None, src_cell=None, src_ws=None, text=None,
                height=None):
    dbs = Side(style='medium', color='A9A9A9')

    if alignment is not None:
        ws[dest_cell].alignment = alignment

    if bg_color is not None:
        ws[dest_cell].fill = bg_color

    if font is not None:
        ws[dest_cell].font = font

    ws[dest_cell].border = Border(left=dbs, right=dbs, top=dbs, bottom=dbs)

    if src_cell is not None and src_ws is not None:
        ws[dest_cell] = src_ws[src_cell].value
    else:
        ws[dest_cell] = text

    if height is not None:
        ws.row_dimensions[ws[dest_cell].row].height = height


if __name__ == '__main__':
    main()
