# import openpyxl
import os
# from openpyxl import load_workbook
# from openpyxl import Workbook
import openpyxl

mytasks = []
ignoredtasks = []

with open('IgnoredTaskCodes.txt') as f:
    ignoredtasks = f.read().splitlines()

ignoredtasks.sort()
print('Task Codes to Ignore:')
print(ignoredtasks)

cwd = os.getcwd()
print(cwd)
# os.chdir("/path/to/folder")
# os.listdir('.')

wb = openpyxl.load_workbook('TimeSheetReportSingleUserGUID.xlsx')
# print(wb2.sheetnames)
# mylength = len(wb2.sheetnames)
# print(mylength)

mysheets = wb.sheetnames

# get tasks from timesheet report
for i in range(len(mysheets)):
    # print(mysheets[i])
    sheet = wb[mysheets[i]]
    # print(sheet['E9'].value)
    j = 9
    while True:
        cell = 'E' + str(j)
        task = sheet[cell].value
        if task == 'Totals:':
            break
        if task not in mytasks:
            mytasks.append(task)
        j += 1

mytasks.sort()
print('\r\nFound the Following Task Codes:')
print(mytasks)

# get hours
taskhours = [0] * len(mytasks)
for i in range(len(mysheets)):
    sheet = wb[mysheets[i]]
    j = 9
    while True:
        taskcell = 'E' + str(j)
        hourcell = 'V' + str(j)
        task = sheet[taskcell].value
        if task == 'Totals:':
            break
        taskindex = mytasks.index(task)
        taskhours[taskindex] += sheet[hourcell].value
        j += 1

totalhours = 0
ignoredhours = 0
print('\r\nTask Code Hours')
for i in range(len(mytasks)):
    totalhours += taskhours[i]
    if mytasks[i] in ignoredtasks:
        ignoredhours += taskhours[i]
    print(mytasks[i] + ' ' + str(taskhours[i]))

if totalhours >= 1000:
    totalspacer = '   '
elif totalhours >= 100:
    totalspacer = '    '
elif totalhours >= 10:
    totalspacer = '     '
else:
    totalspacer = '      '

if ignoredhours >= 1000:
    ignoredspacer = ' '
elif ignoredhours >= 100:
    ignoredspacer = '  '
elif ignoredhours >= 10:
    ignoredspacer = '   '
else:
    ignoredspacer = '    '

workedhours = totalhours - ignoredhours
if workedhours >= 1000:
    workedspacer = ' '
elif workedhours >= 100:
    workedspacer = '  '
elif workedhours >= 10:
    workedspacer = '   '
else:
    workedspacer = '    '

print('\r\nTotal Hours:' + totalspacer + '{:.2f}'.format(totalhours))
print('Ignored Hours:' + ignoredspacer + '{:.2f}'.format(ignoredhours))
print('Working Hours:' + workedspacer + '{:.2f}'.format(workedhours))
print('Average Hours per Week: ' + '{:.2f}'.format(workedhours/52))

print('\r\n')
action = "Z"
while action != "N":
    action = input("Do you want to calculate $/hr? [Y/N]").upper()
    if action not in "YN" or len(action) != 1:
        print("Please choose [Y]es or [N]o")
        continue
    if action == "Y":
        mypay = int(input("How much were you paid? $"))
        print('\r\nTotal $/hr:  ' + '{:.2f}'.format(mypay/totalhours))
        print('Worked $/hr: ' + '{:.2f}'.format(mypay/workedhours))
        break
    else:
        break